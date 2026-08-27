from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from services.monthly_report_drive_service import (
    upload_monthly_report_bytes,
)
from services.report_template_service import (
    NAVY,
    NAVY_2,
    PLN_BLUE,
    SOFT,
    SOFT_BLUE,
    TOTAL_BG,
    TEXT,
    GRID,
    WHITE,
    build_document_id,
    build_modern_header,
    draw_modern_page_decor,
    safe_text,
)


# ==========================================================
# TYPES
# ==========================================================

MonthlyBundle = dict[str, Any]


# ==========================================================
# CONSTANTS
# ==========================================================

MONTH_NAMES: dict[int, str] = {
    1: "JANUARI",
    2: "FEBRUARI",
    3: "MARET",
    4: "APRIL",
    5: "MEI",
    6: "JUNI",
    7: "JULI",
    8: "AGUSTUS",
    9: "SEPTEMBER",
    10: "OKTOBER",
    11: "NOVEMBER",
    12: "DESEMBER",
}

THIN_GRAY = Side(
    style="thin",
    color="B7B7B7",
)


# ==========================================================
# GENERIC HELPERS
# ==========================================================


def _safe_string(
    value: Any,
    default: str = "",
) -> str:
    if value is None:
        return default

    text = str(value).strip()
    return text or default


def _safe_float(
    value: Any,
    default: float = 0.0,
) -> float:
    try:
        if value is None:
            return default

        return float(value)

    except (TypeError, ValueError):
        return default


def _safe_int(
    value: Any,
    default: int = 0,
) -> int:
    try:
        if value is None:
            return default

        return int(float(value))

    except (TypeError, ValueError):
        return default


def _slug(
    value: Any,
) -> str:
    text = _safe_string(
        value,
        "GI",
    )

    cleaned = []

    for char in text:
        if char.isalnum():
            cleaned.append(char)
        elif char in {
            " ",
            "-",
            "_",
        }:
            cleaned.append("_")

    result = "".join(
        cleaned
    )

    while "__" in result:
        result = result.replace(
            "__",
            "_",
        )

    return result.strip(
        "_"
    )


def _period_text(
    report_year: int,
    report_month: int,
) -> str:
    return (
        f"{MONTH_NAMES.get(report_month, str(report_month))} "
        f"{report_year}"
    )


def _format_three_phase_compact(
    row: Any,
    *,
    r_field: str,
    s_field: str,
    t_field: str,
    decimals: int = 0,
) -> str:
    """
    Format ringkas arus 3 phasa untuk PDF.

    Contoh:
        R 102 / S 98 / T 100

    Jika seluruh nilai kosong, hasil kosong.
    """

    r_value = _pdf_cell_text(
        row.get(
            r_field
        ),
        decimals=decimals,
    )

    s_value = _pdf_cell_text(
        row.get(
            s_field
        ),
        decimals=decimals,
    )

    t_value = _pdf_cell_text(
        row.get(
            t_field
        ),
        decimals=decimals,
    )

    if not any(
        (
            r_value,
            s_value,
            t_value,
        )
    ):
        return ""

    return (
        f"R {r_value or '-'} / "
        f"S {s_value or '-'} / "
        f"T {t_value or '-'}"
    )


def _minutes_to_hhmm(
    value: Any,
) -> str:
    total = max(
        0,
        int(
            round(
                _safe_float(
                    value
                )
            )
        ),
    )

    hours = (
        total // 60
    )

    minutes = (
        total % 60
    )

    return (
        f"{hours}:{minutes:02d}"
    )


def _official_guard(
    bundle: MonthlyBundle,
) -> dict[str, Any]:
    report = bundle.get(
        "monthly_report"
    )

    if not isinstance(
        report,
        dict,
    ):
        raise RuntimeError(
            "Header laporan bulanan tidak tersedia."
        )

    status = _safe_string(
        report.get(
            "status"
        )
    ).upper()

    signature_token = _safe_string(
        report.get(
            "signature_token"
        )
    )

    if (
        status != "APPROVED"
        or not signature_token
    ):
        raise RuntimeError(
            "Dokumen resmi hanya dapat dibuat setelah "
            "Verifikasi & e-Sign."
        )

    return report


# ==========================================================
# EXCEL
# ==========================================================


def _excel_style_sheet(
    worksheet: Any,
    *,
    title: str,
    subtitle: str,
    last_column: int,
) -> None:
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )

    worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    worksheet.cell(
        row=1,
        column=1,
    ).font = Font(
        bold=True,
        size=14,
    )

    worksheet.cell(
        row=1,
        column=1,
    ).alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )

    worksheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )

    worksheet.cell(
        row=2,
        column=1,
    ).font = Font(
        bold=True,
        size=11,
    )

    worksheet.cell(
        row=2,
        column=1,
    ).alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[
        1
    ].height = 24

    worksheet.row_dimensions[
        2
    ].height = 20


def _excel_write_table(
    worksheet: Any,
    *,
    dataframe: pd.DataFrame,
    start_row: int = 4,
    blank_rows: int = 0,
    include_total_row: bool = True,
) -> int:
    """
    Menulis tabel Excel secara konsisten.

    - Header selalu dibuat walaupun dataframe kosong.
    - Detail Trip/Lepas dapat mempertahankan baris template kosong.
    - Mengembalikan nomor baris terakhir tabel untuk penempatan e-Sign.
    """

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    total_fill = PatternFill(
        "solid",
        fgColor="EAF3F7",
    )

    border = Border(
        left=THIN_GRAY,
        right=THIN_GRAY,
        top=THIN_GRAY,
        bottom=THIN_GRAY,
    )

    column_count = len(
        dataframe.columns
    )

    if column_count == 0:
        raise ValueError(
            "Struktur kolom laporan Excel tidak tersedia."
        )

    # Header tetap dibuat walaupun tidak ada record.
    for col_index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=start_row,
            column=col_index,
            value=str(
                column
            ),
        )

        cell.font = Font(
            bold=True,
            size=9,
        )

        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[
        start_row
    ].height = 34

    # Data aktual.
    for row_offset, row_values in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        ),
        start=1,
    ):
        excel_row = (
            start_row
            + row_offset
        )

        for col_index, value in enumerate(
            row_values,
            start=1,
        ):
            cell = worksheet.cell(
                row=excel_row,
                column=col_index,
                value=value,
            )

            cell.border = border
            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if col_index
                    <= 2
                    else "left"
                ),
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[
            excel_row
        ].height = 22

    # Baris kosong template.
    required_blank_rows = max(
        0,
        blank_rows
        - len(
            dataframe
        ),
    )

    first_blank_row = (
        start_row
        + len(
            dataframe
        )
        + 1
    )

    for blank_index in range(
        required_blank_rows
    ):
        excel_row = (
            first_blank_row
            + blank_index
        )

        for col_index in range(
            1,
            column_count
            + 1,
        ):
            cell = worksheet.cell(
                row=excel_row,
                column=col_index,
            )

            # Detail laporan menggunakan nomor baris tetap.
            if col_index == 1:
                cell.value = (
                    len(
                        dataframe
                    )
                    + blank_index
                    + 1
                )

            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[
            excel_row
        ].height = 22

    data_body_rows = max(
        len(
            dataframe
        ),
        blank_rows,
    )

    last_body_row = (
        start_row
        + data_body_rows
    )

    last_table_row = last_body_row

    if include_total_row:
        total_row = (
            last_body_row
            + 1
        )

        worksheet.cell(
            row=total_row,
            column=1,
            value="JUMLAH",
        )

        worksheet.cell(
            row=total_row,
            column=1,
        ).font = Font(
            bold=True
        )

        for col_index, column in enumerate(
            dataframe.columns,
            start=1,
        ):
            cell = worksheet.cell(
                row=total_row,
                column=col_index,
            )

            cell.border = border
            cell.fill = total_fill

            if (
                col_index > 1
                and len(
                    dataframe
                )
                > 0
                and pd.api.types.is_numeric_dtype(
                    dataframe[
                        column
                    ]
                )
            ):
                cell.value = (
                    f"=SUM({get_column_letter(col_index)}"
                    f"{start_row + 1}:"
                    f"{get_column_letter(col_index)}"
                    f"{start_row + len(dataframe)})"
                )

                cell.font = Font(
                    bold=True
                )

        worksheet.row_dimensions[
            total_row
        ].height = 22

        last_table_row = (
            total_row
        )

    worksheet.freeze_panes = (
        f"A{start_row + 1}"
    )

    worksheet.auto_filter.ref = (
        f"A{start_row}:"
        f"{get_column_letter(column_count)}"
        f"{max(start_row + 1, last_body_row)}"
    )

    for col_index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        max_length = len(
            str(
                column
            )
        )

        for value in dataframe[
            column
        ].head(
            250
        ):
            max_length = max(
                max_length,
                len(
                    _safe_string(
                        value
                    )
                ),
            )

        worksheet.column_dimensions[
            get_column_letter(
                col_index
            )
        ].width = min(
            max(
                max_length
                + 2,
                10,
            ),
            30,
        )

    return last_table_row



def _excel_add_signature(
    worksheet: Any,
    *,
    signer_name: str,
    signer_position: str,
    verified_at: str,
    start_row: int,
) -> None:
    worksheet.cell(
        row=start_row,
        column=1,
        value="Diverifikasi / Disetujui secara elektronik",
    )

    worksheet.cell(
        row=start_row,
        column=1,
    ).font = Font(
        bold=True,
    )

    worksheet.cell(
        row=start_row + 1,
        column=1,
        value=signer_name,
    )

    worksheet.cell(
        row=start_row + 2,
        column=1,
        value=signer_position,
    )

    worksheet.cell(
        row=start_row + 3,
        column=1,
        value=verified_at,
    )


def _excel_rekap_trip_view(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    columns = {
        "feeder": "FEEDER",
        "feeder_name": "NAMA FEEDER",
        "ocr_inst": "OCR INST",
        "ocr_td": "OCR TD",
        "gfr_inst": "GFR INST",
        "gfr_td": "GFR TD",
        "ufr_uvls_relay": "UFR/UVLS",
        "ols_relay": "OLS",
        "rtn_relay": "RTN",
        "trip_ocr_gfr": "TRIP OCR&GFR",
        "trip_ufr_uvls": "TRIP UFR/UVLS",
        "trip_ols": "TRIP OLS",
        "trip_rtn": "TRIP RTN",
        "total_trip_event": "TOTAL TRIP",
        "menit_ocr_gfr": "MENIT OCR&GFR",
        "menit_ufr_uvls": "MENIT UFR/UVLS",
        "menit_ols": "MENIT OLS",
        "menit_rtn": "MENIT RTN",
        "total_menit": "TOTAL MENIT",
        "kwh_ocr_gfr": "KWH OCR&GFR",
        "kwh_ufr_uvls": "KWH UFR/UVLS",
        "kwh_ols": "KWH OLS",
        "kwh_rtn": "KWH RTN",
        "total_kwh": "TOTAL KWH",
    }

    existing = [
        column
        for column in columns
        if column in dataframe.columns
    ]

    return dataframe[
        existing
    ].rename(
        columns=columns
    )


def _excel_rekap_lepas_view(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    columns = {
        "feeder": "FEEDER",
        "feeder_name": "NAMA FEEDER",
        "lepas_har": "HAR",
        "lepas_defisit": "DEFISIT",
        "lepas_ulp": "ULP",
        "lepas_emergency_upt": "EMERGENCY UPT",
        "lepas_emergency_ulp": "EMERGENCY ULP",
        "lepas_blackout": "BLACKOUT",
        "lepas_lainnya": "LAINNYA",
        "jumlah_lepas": "JUMLAH LEPAS",
        "menit_har": "MENIT HAR",
        "menit_defisit": "MENIT DEFISIT",
        "menit_ulp": "MENIT ULP",
        "menit_emergency_upt": "MENIT EMERGENCY UPT",
        "menit_emergency_ulp": "MENIT EMERGENCY ULP",
        "menit_blackout": "MENIT BLACKOUT",
        "menit_lainnya": "MENIT LAINNYA",
        "total_menit": "TOTAL MENIT",
        "kwh_har": "KWH HAR",
        "kwh_defisit": "KWH DEFISIT",
        "kwh_ulp": "KWH ULP",
        "kwh_emergency_upt": "KWH EMERGENCY UPT",
        "kwh_emergency_ulp": "KWH EMERGENCY ULP",
        "kwh_blackout": "KWH BLACKOUT",
        "kwh_lainnya": "KWH LAINNYA",
        "total_kwh": "TOTAL KWH",
    }

    existing = [
        column
        for column in columns
        if column in dataframe.columns
    ]

    return dataframe[
        existing
    ].rename(
        columns=columns
    )


def _excel_detail_trip_view(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    columns = {
        "no": "NO",
        "nama_penyulang": "NAMA PENYULANG",
        "kondisi": "KONDISI",
        "tgl": "TGL",
        "pkl": "PKL",

        # Arus beban sebelum gangguan - 3 phasa
        "amp_r": "ARUS SEBELUM R (A)",
        "amp_s": "ARUS SEBELUM S (A)",
        "amp_t": "ARUS SEBELUM T (A)",

        "kv": "KV",

        # Arus gangguan / fault current
        "r": "ARUS GGN R (A)",
        "s": "ARUS GGN S (A)",
        "t": "ARUS GGN T (A)",
        "n": "ARUS GGN N (A)",

        "pemulihan_kondisi": "PEMULIHAN",
        "pemulihan_tgl": "TGL PULIH",
        "pemulihan_pkl": "PKL PULIH",

        # Arus beban setelah pemulihan - 3 phasa
        "amp_after_r": "ARUS SETELAH R (A)",
        "amp_after_s": "ARUS SETELAH S (A)",
        "amp_after_t": "ARUS SETELAH T (A)",

        # Pemulihan beban / manuver
        "supply_status_name": "STATUS SUPLAI",
        "supply_restored_date": "TGL MULAI TERSUPLAI",
        "supply_restored_time": "PKL MULAI TERSUPLAI",

        "maneuvered_r": "TERMANUVER R (A)",
        "maneuvered_s": "TERMANUVER S (A)",
        "maneuvered_t": "TERMANUVER T (A)",

        "remaining_r": "SISA R (A)",
        "remaining_s": "SISA S (A)",
        "remaining_t": "SISA T (A)",

        "final_supply_normalized": "BEBAN NORMAL",
        "final_supply_normalization_date": "TGL NORMALISASI BEBAN",
        "final_supply_normalization_time": "PKL NORMALISASI BEBAN",

        "menit": "MENIT",
        "jlh_kwh": "KWH PADAM",
        "annunciator": "ANNUNCIATOR",
        "indikasi_name": "RELE YANG BEKERJA",
        "phasa": "PHASA",
        "penyebab_kejadian": "PENYEBAB KEJADIAN",
        "keterangan": "KETERANGAN",
        "operator_bertugas": "OPERATOR",
        "dispatcher_up2d": "DISPATCHER UP2D",
        "diinput_oleh": "DIINPUT OLEH",
    }

    # Reindex menjaga struktur kolom meskipun dataframe benar-benar kosong.
    return (
        dataframe
        .reindex(
            columns=list(
                columns.keys()
            )
        )
        .rename(
            columns=columns
        )
    )


def _excel_detail_lepas_view(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    columns = {
        "no": "NO",
        "nama_penyulang": "NAMA PENYULANG",
        "kondisi": "KONDISI",
        "tgl": "TGL",
        "pkl": "PKL",

        # Arus beban sebelum manuver/lepas - 3 phasa
        "amp_r": "ARUS SEBELUM R (A)",
        "amp_s": "ARUS SEBELUM S (A)",
        "amp_t": "ARUS SEBELUM T (A)",

        "kv": "KV",

        "pemulihan_kondisi": "PEMULIHAN",
        "pemulihan_tgl": "TGL PULIH",
        "pemulihan_pkl": "PKL PULIH",

        # Arus beban setelah normalisasi - 3 phasa
        "amp_after_r": "ARUS SETELAH R (A)",
        "amp_after_s": "ARUS SETELAH S (A)",
        "amp_after_t": "ARUS SETELAH T (A)",

        # Pemulihan beban / manuver
        "supply_status_name": "STATUS SUPLAI",
        "supply_restored_date": "TGL MULAI TERSUPLAI",
        "supply_restored_time": "PKL MULAI TERSUPLAI",

        "maneuvered_r": "TERMANUVER R (A)",
        "maneuvered_s": "TERMANUVER S (A)",
        "maneuvered_t": "TERMANUVER T (A)",

        "remaining_r": "SISA R (A)",
        "remaining_s": "SISA S (A)",
        "remaining_t": "SISA T (A)",

        "final_supply_normalized": "BEBAN NORMAL",
        "final_supply_normalization_date": "TGL NORMALISASI BEBAN",
        "final_supply_normalization_time": "PKL NORMALISASI BEBAN",

        "menit": "MENIT",
        "jlh_kwh": "KWH PADAM",
        "penyebab_kejadian": "PENYEBAB KEJADIAN",
        "kategori_lepas": "KATEGORI",
        "keterangan": "KETERANGAN",
        "operator_bertugas": "OPERATOR",
        "dispatcher_up2d": "DISPATCHER UP2D",
        "diinput_oleh": "DIINPUT OLEH",
    }

    # Reindex menjaga struktur kolom meskipun dataframe benar-benar kosong.
    return (
        dataframe
        .reindex(
            columns=list(
                columns.keys()
            )
        )
        .rename(
            columns=columns
        )
    )


def build_monthly_excel_bytes(
    *,
    bundle: MonthlyBundle,
    report_year: int,
    report_month: int,
    gi_name: str,
) -> bytes:
    report = _official_guard(
        bundle
    )

    signer_name = _safe_string(
        report.get(
            "signer_name"
        ),
        "-",
    )

    signer_position = _safe_string(
        report.get(
            "signer_position"
        ),
        _safe_string(
            report.get(
                "verified_role"
            ),
            "-",
        ),
    )

    verified_at = _safe_string(
        report.get(
            "verified_at"
        ),
        "-",
    )

    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(
            default_sheet
        )

    sheet_specs = [
        (
            "Rekap Trip",
            "REKAPITULASI DATA JUMLAH TRIP PMT 20 KV",
            _excel_rekap_trip_view(
                bundle[
                    "rekap_trip"
                ]
            ),
            0,
            True,
        ),
        (
            "Rekap Lepas",
            "REKAPITULASI DATA JUMLAH LEPAS PMT 20 KV",
            _excel_rekap_lepas_view(
                bundle[
                    "rekap_lepas"
                ]
            ),
            0,
            True,
        ),
        (
            "Detail Trip",
            "LAPORAN KONDISI TRIP PENYULANG 20 KV / BUSTIE",
            _excel_detail_trip_view(
                bundle[
                    "detail_trip"
                ]
            ),
            12,
            False,
        ),
        (
            "Detail Lepas",
            "LAPORAN KONDISI LEPAS PENYULANG 20 KV / BUSTIE",
            _excel_detail_lepas_view(
                bundle[
                    "detail_lepas"
                ]
            ),
            12,
            False,
        ),
    ]

    subtitle = (
        f"BULAN : {_period_text(report_year, report_month)}"
        f" | GI : {gi_name}"
    )

    for (
        sheet_name,
        title,
        dataframe,
        blank_rows,
        include_total_row,
    ) in sheet_specs:
        worksheet = workbook.create_sheet(
            title=sheet_name
        )

        last_column = max(
            1,
            len(
                dataframe.columns
            ),
        )

        _excel_style_sheet(
            worksheet,
            title=title,
            subtitle=subtitle,
            last_column=last_column,
        )

        last_table_row = _excel_write_table(
            worksheet,
            dataframe=dataframe,
            start_row=4,
            blank_rows=blank_rows,
            include_total_row=include_total_row,
        )

        signature_row = (
            last_table_row
            + 4
        )

        _excel_add_signature(
            worksheet,
            signer_name=signer_name,
            signer_position=signer_position,
            verified_at=verified_at,
            start_row=signature_row,
        )

        worksheet.page_setup.orientation = (
            "landscape"
        )

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(
        0
    )

    return output.getvalue()

    # ==========================================================
# PDF V2 - FIXED PROFESSIONAL TEMPLATE
# ==========================================================


def _pdf_cell_text(
    value: Any,
    *,
    decimals: int | None = None,
) -> str:
    if value is None:
        return ""

    if isinstance(
        value,
        float,
    ):
        try:
            if pd.isna(
                value
            ):
                return ""
        except Exception:
            pass

    text = safe_text(
        value
    )

    if not text:
        return ""

    if decimals is not None:
        try:
            number = float(
                value
            )

            if pd.isna(
                number
            ):
                return ""

            return (
                f"{number:.{decimals}f}"
                .replace(
                    ".",
                    ",",
                )
            )

        except Exception:
            return text

    return text


def _pdf_minutes_hhmm(
    value: Any,
) -> str:
    try:
        if value is None:
            return ""

        minutes = float(
            value
        )

        if pd.isna(
            minutes
        ):
            return ""

    except Exception:
        return ""

    total = max(
        0,
        int(
            round(
                minutes
            )
        ),
    )

    return (
        f"{total // 60}:"
        f"{total % 60:02d}"
    )


def _professional_table_style(
    *,
    header_rows: int,
    total_row_index: int | None = None,
    font_size: float = 6.2,
) -> TableStyle:
    commands: list[tuple[Any, ...]] = [
        ("GRID", (0, 0), (-1, -1), 0.3, GRID),
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, header_rows - 1), SOFT_BLUE),
        ("TEXTCOLOR", (0, 1), (-1, header_rows - 1), TEXT),
        ("FONTNAME", (0, 1), (-1, header_rows - 1), "Helvetica-Bold"),
        ("FONTNAME", (0, header_rows), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("TEXTCOLOR", (0, header_rows), (-1, -1), TEXT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("VALIGN", (0, 0), (-1, header_rows - 1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.4),
        ("TOPPADDING", (0, 0), (-1, -1), 2.0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.0),
        ("LINEBELOW", (0, header_rows - 1), (-1, header_rows - 1), 0.55, NAVY_2),
    ]

    if total_row_index is not None:
        commands.extend([
            ("BACKGROUND", (0, total_row_index), (-1, total_row_index), TOTAL_BG),
            ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),
            ("LINEABOVE", (0, total_row_index), (-1, total_row_index), 0.8, NAVY_2),
        ])

    # Soft zebra striping for body rows.
    if total_row_index is not None:
        body_end = total_row_index - 1
    else:
        body_end = -1

    for row_index in range(header_rows, body_end + 1 if body_end >= 0 else header_rows):
        if (row_index - header_rows) % 2 == 1:
            commands.append(("BACKGROUND", (0, row_index), (-1, row_index), SOFT))

    return TableStyle(commands)



def _header_cell(text: str, *, inverse: bool = False, size: float = 5.7) -> Paragraph:
    color = WHITE if inverse else TEXT
    return Paragraph(
        text,
        ParagraphStyle(
            f"hdr_{id(text)}_{inverse}",
            parent=getSampleStyleSheet()["Normal"],
            fontName="Helvetica-Bold",
            fontSize=size,
            leading=size * 1.10,
            alignment=TA_CENTER,
            textColor=color,
        ),
    )


def _modernize_header_rows(header: list[list[Any]]) -> list[list[Any]]:
    modern: list[list[Any]] = []
    for row_index, row in enumerate(header):
        converted: list[Any] = []
        for value in row:
            if isinstance(value, str) and value:
                converted.append(
                    _header_cell(
                        value.replace("\\n", "<br/>"),
                        inverse=(row_index == 0),
                        size=5.6 if row_index == 0 else 5.35,
                    )
                )
            else:
                converted.append(value)
        modern.append(converted)
    return modern


def _trip_recap_pdf_table(
    dataframe: pd.DataFrame,
) -> Table:
    header = [
        [
            "NO",
            "FEEDER",
            "RELE YANG BEKERJA",
            "",
            "",
            "",
            "",
            "",
            "",
            "JUMLAH TRIP",
            "",
            "",
            "",
            "JUMLAH WAKTU / MENIT",
            "",
            "",
            "",
            "KWH PADAM",
            "",
            "",
            "",
        ],
        [
            "",
            "",
            "OCR",
            "",
            "GFR",
            "",
            "UFR/UVLS",
            "OLS",
            "RTN",
            "OCR & GFR",
            "UFR/UVLS",
            "OLS",
            "RTN",
            "OCR & GFR",
            "UFR/UVLS",
            "OLS",
            "RTN",
            "OCR & GFR",
            "UFR/UVLS",
            "OLS",
            "RTN",
        ],
        [
            "",
            "",
            "INST",
            "TD",
            "INST",
            "TD",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]

    body: list[list[Any]] = []

    for row_number, (_, row) in enumerate(
        dataframe.reset_index(
            drop=True
        ).iterrows(),
        start=1,
    ):
        feeder = _pdf_cell_text(
            row.get(
                "feeder"
            )
        )

        alias = _pdf_cell_text(
            row.get(
                "feeder_alias"
            )
        )

        if alias:
            feeder = (
                f"{feeder} ({alias})"
            )

        body.append(
            [
                row_number,
                feeder,
                _pdf_cell_text(row.get("ocr_inst")),
                _pdf_cell_text(row.get("ocr_td")),
                _pdf_cell_text(row.get("gfr_inst")),
                _pdf_cell_text(row.get("gfr_td")),
                _pdf_cell_text(row.get("ufr_uvls_relay")),
                _pdf_cell_text(row.get("ols_relay")),
                _pdf_cell_text(row.get("rtn_relay")),
                _pdf_cell_text(row.get("trip_ocr_gfr")),
                _pdf_cell_text(row.get("trip_ufr_uvls")),
                _pdf_cell_text(row.get("trip_ols")),
                _pdf_cell_text(row.get("trip_rtn")),
                _pdf_cell_text(row.get("menit_ocr_gfr"), decimals=1),
                _pdf_cell_text(row.get("menit_ufr_uvls"), decimals=1),
                _pdf_cell_text(row.get("menit_ols"), decimals=1),
                _pdf_cell_text(row.get("menit_rtn"), decimals=1),
                _pdf_cell_text(row.get("kwh_ocr_gfr"), decimals=2),
                _pdf_cell_text(row.get("kwh_ufr_uvls"), decimals=2),
                _pdf_cell_text(row.get("kwh_ols"), decimals=2),
                _pdf_cell_text(row.get("kwh_rtn"), decimals=2),
            ]
        )

    total_row: list[Any] = [
        "",
        "JUMLAH",
    ]

    total_fields = [
        "ocr_inst",
        "ocr_td",
        "gfr_inst",
        "gfr_td",
        "ufr_uvls_relay",
        "ols_relay",
        "rtn_relay",
        "trip_ocr_gfr",
        "trip_ufr_uvls",
        "trip_ols",
        "trip_rtn",
        "menit_ocr_gfr",
        "menit_ufr_uvls",
        "menit_ols",
        "menit_rtn",
        "kwh_ocr_gfr",
        "kwh_ufr_uvls",
        "kwh_ols",
        "kwh_rtn",
    ]

    for field in total_fields:
        value = (
            pd.to_numeric(
                dataframe.get(
                    field,
                    pd.Series(
                        dtype=float
                    ),
                ),
                errors="coerce",
            )
            .fillna(
                0
            )
            .sum()
        )

        decimals = (
            2
            if field.startswith(
                "kwh_"
            )
            else (
                1
                if field.startswith(
                    "menit_"
                )
                else 0
            )
        )

        total_row.append(
            _pdf_cell_text(
                value,
                decimals=decimals,
            )
        )

    data = (
        _modernize_header_rows(header)
        + body
        + [
            total_row
        ]
    )

    widths_mm = [
        7,
        36,
        10,
        10,
        10,
        10,
        12,
        10,
        10,
        15,
        12,
        10,
        10,
        15,
        12,
        10,
        10,
        15,
        12,
        10,
        10,
    ]

    row_heights: list[Any] = [
        8.5 * mm,
        8.0 * mm,
        7.0 * mm,
    ]

    row_heights.extend(
        [
            6.0 * mm
            for _ in range(
                max(
                    0,
                    len(data) - 3,
                )
            )
        ]
    )

    table = Table(
        data,
        colWidths=[
            width * mm
            for width in widths_mm
        ],
        rowHeights=row_heights,
        repeatRows=3,
        hAlign="CENTER",
    )

    style = _professional_table_style(
        header_rows=3,
        total_row_index=len(
            data
        ) - 1,
        font_size=6.15,
    )

    style.add(
        "SPAN",
        (0, 0),
        (0, 2),
    )

    style.add(
        "SPAN",
        (1, 0),
        (1, 2),
    )

    for column in (
        0,
        1,
    ):
        style.add(
            "BACKGROUND",
            (column, 0),
            (column, 2),
            NAVY,
        )
        style.add(
            "TEXTCOLOR",
            (column, 0),
            (column, 2),
            WHITE,
        )

    style.add(
        "SPAN",
        (2, 0),
        (8, 0),
    )

    style.add(
        "SPAN",
        (9, 0),
        (12, 0),
    )

    style.add(
        "SPAN",
        (13, 0),
        (16, 0),
    )

    style.add(
        "SPAN",
        (17, 0),
        (20, 0),
    )

    style.add(
        "SPAN",
        (2, 1),
        (3, 1),
    )

    style.add(
        "SPAN",
        (4, 1),
        (5, 1),
    )

    for column in range(
        6,
        21,
    ):
        if column not in {
            2,
            3,
            4,
            5,
        }:
            style.add(
                "SPAN",
                (column, 1),
                (column, 2),
            )

    table.setStyle(
        style
    )

    return table


def _lepas_recap_pdf_table(
    dataframe: pd.DataFrame,
) -> Table:
    header = [
        [
            "NO",
            "FEEDER",
            "DATA LEPAS",
            "",
            "",
            "",
            "",
            "",
            "",
            "JUMLAH LEPAS",
            "JUMLAH WAKTU / MENIT",
            "",
            "",
            "",
            "",
            "",
            "",
            "TOTAL WAKTU",
            "KWH PADAM",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "",
            "",
            "HAR",
            "DEFISIT",
            "ULP",
            "EMERGENCY<br/>UPT",
            "EMERGENCY<br/>ULP",
            "BLACKOUT",
            "LAINNYA",
            "",
            "HAR",
            "DEFISIT",
            "ULP",
            "EMERGENCY<br/>UPT",
            "EMERGENCY<br/>ULP",
            "BLACKOUT",
            "LAINNYA",
            "",
            "HAR",
            "DEFISIT",
            "ULP",
            "EMERGENCY<br/>UPT",
            "EMERGENCY<br/>ULP",
            "BLACKOUT",
            "LAINNYA",
        ],
    ]

    body: list[list[Any]] = []

    for row_number, (_, row) in enumerate(
        dataframe.reset_index(
            drop=True
        ).iterrows(),
        start=1,
    ):
        feeder = _pdf_cell_text(
            row.get(
                "feeder"
            )
        )

        alias = _pdf_cell_text(
            row.get(
                "feeder_alias"
            )
        )

        if alias:
            feeder = (
                f"{feeder} ({alias})"
            )

        body.append(
            [
                row_number,
                feeder,
                _pdf_cell_text(row.get("lepas_har")),
                _pdf_cell_text(row.get("lepas_defisit")),
                _pdf_cell_text(row.get("lepas_ulp")),
                _pdf_cell_text(row.get("lepas_emergency_upt")),
                _pdf_cell_text(row.get("lepas_emergency_ulp")),
                _pdf_cell_text(row.get("lepas_blackout")),
                _pdf_cell_text(row.get("lepas_lainnya")),
                _pdf_cell_text(row.get("jumlah_lepas")),
                _pdf_cell_text(row.get("menit_har"), decimals=1),
                _pdf_cell_text(row.get("menit_defisit"), decimals=1),
                _pdf_cell_text(row.get("menit_ulp"), decimals=1),
                _pdf_cell_text(row.get("menit_emergency_upt"), decimals=1),
                _pdf_cell_text(row.get("menit_emergency_ulp"), decimals=1),
                _pdf_cell_text(row.get("menit_blackout"), decimals=1),
                _pdf_cell_text(row.get("menit_lainnya"), decimals=1),
                _pdf_cell_text(row.get("total_menit"), decimals=1),
                _pdf_cell_text(row.get("kwh_har"), decimals=2),
                _pdf_cell_text(row.get("kwh_defisit"), decimals=2),
                _pdf_cell_text(row.get("kwh_ulp"), decimals=2),
                _pdf_cell_text(row.get("kwh_emergency_upt"), decimals=2),
                _pdf_cell_text(row.get("kwh_emergency_ulp"), decimals=2),
                _pdf_cell_text(row.get("kwh_blackout"), decimals=2),
                _pdf_cell_text(row.get("kwh_lainnya"), decimals=2),
            ]
        )

    total_row: list[Any] = [
        "",
        "JUMLAH",
    ]

    for field in [
        "lepas_har",
        "lepas_defisit",
        "lepas_ulp",
        "lepas_emergency_upt",
        "lepas_emergency_ulp",
        "lepas_blackout",
        "lepas_lainnya",
        "jumlah_lepas",
        "menit_har",
        "menit_defisit",
        "menit_ulp",
        "menit_emergency_upt",
        "menit_emergency_ulp",
        "menit_blackout",
        "menit_lainnya",
        "total_menit",
        "kwh_har",
        "kwh_defisit",
        "kwh_ulp",
        "kwh_emergency_upt",
        "kwh_emergency_ulp",
        "kwh_blackout",
        "kwh_lainnya",
    ]:
        value = (
            pd.to_numeric(
                dataframe.get(
                    field,
                    pd.Series(
                        dtype=float
                    ),
                ),
                errors="coerce",
            )
            .fillna(
                0
            )
            .sum()
        )

        decimals = (
            2
            if field.startswith(
                "kwh_"
            )
            else (
                1
                if field.startswith(
                    "menit_"
                )
                or field
                == "total_menit"
                else 0
            )
        )

        total_row.append(
            _pdf_cell_text(
                value,
                decimals=decimals,
            )
        )

    data = (
        _modernize_header_rows(header)
        + body
        + [
            total_row
        ]
    )

    widths_mm = [
        7,
        34,
        8,
        9,
        8,
        12,
        12,
        9,
        9,
        10,
        8,
        9,
        8,
        12,
        12,
        9,
        9,
        11,
        9,
        9,
        9,
        12,
        12,
        9,
        9,
    ]

    row_heights: list[Any] = [
        10.0 * mm,
        13.5 * mm,
    ]

    row_heights.extend(
        [
            6.0 * mm
            for _ in range(
                max(
                    0,
                    len(data) - 2,
                )
            )
        ]
    )

    table = Table(
        data,
        colWidths=[
            width * mm
            for width in widths_mm
        ],
        rowHeights=row_heights,
        repeatRows=2,
        hAlign="CENTER",
    )

    style = _professional_table_style(
        header_rows=2,
        total_row_index=len(
            data
        ) - 1,
        font_size=5.45,
    )

    style.add(
        "SPAN",
        (0, 0),
        (0, 1),
    )

    style.add(
        "SPAN",
        (1, 0),
        (1, 1),
    )

    style.add(
        "SPAN",
        (2, 0),
        (8, 0),
    )

    style.add(
        "SPAN",
        (9, 0),
        (9, 1),
    )

    style.add(
        "SPAN",
        (10, 0),
        (16, 0),
    )

    style.add(
        "SPAN",
        (17, 0),
        (17, 1),
    )

    for column in (
        0,
        1,
        9,
        17,
    ):
        style.add(
            "BACKGROUND",
            (column, 0),
            (column, 1),
            NAVY,
        )
        style.add(
            "TEXTCOLOR",
            (column, 0),
            (column, 1),
            WHITE,
        )

    style.add(
        "SPAN",
        (18, 0),
        (24, 0),
    )

    table.setStyle(
        style
    )

    return table


def _detail_pdf_table(
    dataframe: pd.DataFrame,
    *,
    is_trip: bool,
) -> Table:
    """
    Detail Trip / Lepas pada landscape A4.

    Arus beban sebelum, setelah, termanuver, dan sisa beban
    ditampilkan ringkas sebagai R/S/T dalam satu kolom masing-masing.
    Arus gangguan R/S/T/N tetap dipisahkan.

    Kolom MENIT dipertahankan sebagai bagian laporan resmi.
    """

    header = [
        [
            "NO",
            "NAMA PENYULANG",
            "URAIAN KEJADIAN",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "KWH PADAM",
            "ANNUNCIATOR",
            "RELE YANG BEKERJA",
            "",
            "PENYEBAB KEJADIAN",
            "KETERANGAN",
        ],
        [
            "",
            "",
            "KONDISI",
            "TGL",
            "PKL",
            "ARUS BEBAN<br/>SEBELUM (A)",
            "KV",
            "ARUS GGN (A)",
            "",
            "",
            "",
            "PEMULIHAN",
            "",
            "",
            "ARUS BEBAN<br/>SETELAH (A)",
            "BEBAN<br/>TERMANUVER (A)",
            "SISA BEBAN<br/>(A)",
            "MENIT",
            "JLH KWH",
            "",
            "INDIKASI",
            "PHASA",
            "",
            "",
        ],
        [
            "",
            "",
            "",
            "",
            "",
            "R / S / T",
            "",
            "R",
            "S",
            "T",
            "N",
            "KONDISI",
            "TGL",
            "PKL",
            "R / S / T",
            "R / S / T",
            "R / S / T",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]

    body: list[list[Any]] = []

    for row_number, (_, row) in enumerate(
        dataframe.reset_index(
            drop=True
        ).iterrows(),
        start=1,
    ):
        current_before = _format_three_phase_compact(
            row,
            r_field="amp_r",
            s_field="amp_s",
            t_field="amp_t",
            decimals=0,
        )

        current_after = _format_three_phase_compact(
            row,
            r_field="amp_after_r",
            s_field="amp_after_s",
            t_field="amp_after_t",
            decimals=0,
        )

        maneuvered_current = _format_three_phase_compact(
            row,
            r_field="maneuvered_r",
            s_field="maneuvered_s",
            t_field="maneuvered_t",
            decimals=0,
        )

        remaining_current = _format_three_phase_compact(
            row,
            r_field="remaining_r",
            s_field="remaining_s",
            t_field="remaining_t",
            decimals=0,
        )

        body.append(
            [
                row_number,
                _pdf_cell_text(
                    row.get(
                        "nama_penyulang"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "kondisi"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "tgl"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "pkl"
                    )
                ),
                current_before,
                _pdf_cell_text(
                    row.get(
                        "kv"
                    ),
                    decimals=1,
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "r"
                        ),
                        decimals=0,
                    )
                    if is_trip
                    else ""
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "s"
                        ),
                        decimals=0,
                    )
                    if is_trip
                    else ""
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "t"
                        ),
                        decimals=0,
                    )
                    if is_trip
                    else ""
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "n"
                        ),
                        decimals=0,
                    )
                    if is_trip
                    else ""
                ),
                _pdf_cell_text(
                    row.get(
                        "pemulihan_kondisi"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "pemulihan_tgl"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "pemulihan_pkl"
                    )
                ),
                current_after,
                maneuvered_current,
                remaining_current,
                _pdf_minutes_hhmm(
                    row.get(
                        "menit"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "jlh_kwh"
                    ),
                    decimals=2,
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "annunciator"
                        )
                    )
                    if is_trip
                    else ""
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "indikasi_name"
                        )
                    )
                    if is_trip
                    else ""
                ),
                (
                    _pdf_cell_text(
                        row.get(
                            "phasa"
                        )
                    )
                    if is_trip
                    else ""
                ),
                _pdf_cell_text(
                    row.get(
                        "penyebab_kejadian"
                    )
                ),
                _pdf_cell_text(
                    row.get(
                        "keterangan"
                    )
                ),
            ]
        )

    target_rows = max(
        12,
        len(
            body
        ),
    )

    while len(
        body
    ) < target_rows:
        body.append(
            [
                len(
                    body
                )
                + 1,
                *[
                    ""
                    for _ in range(
                        23
                    )
                ],
            ]
        )

    data = (
        _modernize_header_rows(
            header
        )
        + body
    )

    # Total 279 mm, masih aman untuk landscape A4.
    widths_mm = [
        5,   # 0  NO
        20,  # 1  Nama penyulang
        9,   # 2  Kondisi
        7,   # 3  Tgl
        8,   # 4  Pkl
        22,  # 5  Arus sebelum R/S/T
        6,   # 6  kV
        6,   # 7  fault R
        6,   # 8  fault S
        6,   # 9  fault T
        6,   # 10 fault N
        9,   # 11 Pemulihan kondisi
        7,   # 12 Tgl pulih
        8,   # 13 Pkl pulih
        22,  # 14 Arus setelah R/S/T
        20,  # 15 Termanuver R/S/T
        20,  # 16 Sisa R/S/T
        8,   # 17 Menit
        9,   # 18 kWh
        15,  # 19 Annunciator
        12,  # 20 Rele
        7,   # 21 Phasa
        18,  # 22 Penyebab
        23,  # 23 Keterangan
    ]

    row_heights: list[Any] = [
        10.0 * mm,
        10.0 * mm,
        8.0 * mm,
    ]

    row_heights.extend(
        [
            6.4 * mm
            for _ in range(
                max(
                    0,
                    len(data)
                    - 3,
                )
            )
        ]
    )

    table = Table(
        data,
        colWidths=[
            width * mm
            for width in widths_mm
        ],
        rowHeights=row_heights,
        repeatRows=3,
        hAlign="CENTER",
    )

    style = _professional_table_style(
        header_rows=3,
        font_size=4.45,
    )

    for command in [
        ("SPAN", (0, 0), (0, 2)),
        ("SPAN", (1, 0), (1, 2)),
        ("SPAN", (2, 0), (17, 0)),
        ("SPAN", (18, 0), (18, 2)),
        ("SPAN", (19, 0), (19, 2)),
        ("SPAN", (20, 0), (21, 0)),
        ("SPAN", (22, 0), (22, 2)),
        ("SPAN", (23, 0), (23, 2)),

        ("SPAN", (2, 1), (2, 2)),
        ("SPAN", (3, 1), (3, 2)),
        ("SPAN", (4, 1), (4, 2)),
        ("SPAN", (6, 1), (6, 2)),
        ("SPAN", (7, 1), (10, 1)),
        ("SPAN", (11, 1), (13, 1)),
        ("SPAN", (17, 1), (17, 2)),
        ("SPAN", (20, 1), (20, 2)),
        ("SPAN", (21, 1), (21, 2)),
    ]:
        style.add(
            *command
        )

    for column in (
        0,
        1,
        18,
        19,
        22,
        23,
    ):
        style.add(
            "BACKGROUND",
            (column, 0),
            (column, 2),
            NAVY,
        )

        style.add(
            "TEXTCOLOR",
            (column, 0),
            (column, 2),
            WHITE,
        )

    # Arus beban dan manuver diberi aksen ringan.
    for column in (
        5,
        14,
        15,
        16,
    ):
        style.add(
            "BACKGROUND",
            (column, 1),
            (column, 2),
            SOFT_BLUE,
        )

    # Untuk laporan Lepas, arus gangguan dan proteksi tidak digunakan.
    if not is_trip:
        style.add(
            "BACKGROUND",
            (7, 3),
            (10, -1),
            colors.HexColor(
                "#E5E5E5"
            ),
        )

        style.add(
            "BACKGROUND",
            (19, 3),
            (21, -1),
            colors.HexColor(
                "#E5E5E5"
            ),
        )

    table.setStyle(
        style
    )

    return table

def build_monthly_pdf_bytes(
    *,
    bundle: MonthlyBundle,
    report_year: int,
    report_month: int,
    gi_name: str,
    gi_flc: str = "GI",
) -> bytes:
    report = _official_guard(bundle)

    monthly_report_id = safe_text(
        report.get("monthly_report_id"),
        "-",
    )

    document_id = build_document_id(
        report_year=report_year,
        report_month=report_month,
        gi_flc=gi_flc,
        monthly_report_id=monthly_report_id,
    )

    period_text = _period_text(
        report_year,
        report_month,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=7.5 * mm,
        rightMargin=7.5 * mm,
        topMargin=10 * mm,
        bottomMargin=31 * mm,
        title=f"{document_id} - Laporan Bulanan {gi_name}",
        author="Laporan Gangguan Penyulang",
    )

    story: list[Any] = []

    rekap_trip = bundle["rekap_trip"]
    rekap_lepas = bundle["rekap_lepas"]
    detail_trip = bundle["detail_trip"]
    detail_lepas = bundle["detail_lepas"]

    trip_total = int(pd.to_numeric(rekap_trip.get("total_trip_event", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    trip_minutes = float(pd.to_numeric(rekap_trip.get("total_menit", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    trip_kwh = float(pd.to_numeric(rekap_trip.get("total_kwh", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    trip_feeders = int((pd.to_numeric(rekap_trip.get("total_trip_event", pd.Series(dtype=float)), errors="coerce").fillna(0) > 0).sum())

    lepas_total = int(pd.to_numeric(rekap_lepas.get("jumlah_lepas", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    lepas_minutes = float(pd.to_numeric(rekap_lepas.get("total_menit", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    lepas_kwh = float(pd.to_numeric(rekap_lepas.get("total_kwh", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    lepas_feeders = int((pd.to_numeric(rekap_lepas.get("jumlah_lepas", pd.Series(dtype=float)), errors="coerce").fillna(0) > 0).sum())

    sections = [
        (
            "REKAPITULASI DATA JUMLAH TRIP PMT 20 kV",
            _trip_recap_pdf_table(rekap_trip),
        ),
        (
            "REKAPITULASI DATA JUMLAH LEPAS PMT 20 kV",
            _lepas_recap_pdf_table(rekap_lepas),
        ),
        (
            "LAPORAN KONDISI TRIP PENYULANG 20 kV / BUSTIE",
            _detail_pdf_table(detail_trip, is_trip=True),
        ),
        (
            "LAPORAN KONDISI LEPAS PENYULANG 20 kV / BUSTIE",
            _detail_pdf_table(detail_lepas, is_trip=False),
        ),
    ]

    for page_index, (title, table) in enumerate(sections):
        if page_index > 0:
            story.append(PageBreak())

        story.append(
            build_modern_header(
                gi_name=gi_name,
                report_title=title,
                period_text=period_text,
                document_id=document_id,
            )
        )

        story.append(
            Spacer(
                1,
                5.5 * mm,
            )
        )

        story.append(
            table
        )

    def _on_page(canvas: Any, doc: Any) -> None:
        draw_modern_page_decor(
            canvas,
            doc,
            report=report,
            gi_name=gi_name,
            gi_flc=gi_flc,
            report_year=report_year,
            report_month=report_month,
        )

    document.build(
        story,
        onFirstPage=_on_page,
        onLaterPages=_on_page,
    )

    output.seek(0)
    return output.getvalue()


# ==========================================================
# EXPORT + GOOGLE DRIVE
# ==========================================================


def generate_and_archive_official_report(
    *,
    bundle: MonthlyBundle,
    report_year: int,
    report_month: int,
    gi_flc: str,
    gi_name: str,
) -> dict[str, Any]:
    report = _official_guard(
        bundle
    )

    monthly_report_id = _safe_string(
        report.get(
            "monthly_report_id"
        )
    )

    if not monthly_report_id:
        raise RuntimeError(
            "monthly_report_id tidak tersedia."
        )

    period_code = (
        f"{report_year}-"
        f"{report_month:02d}"
    )

    gi_slug = _slug(
        gi_name
    )

    pdf_name = (
        f"Laporan_Bulanan_"
        f"{gi_slug}_"
        f"{period_code}.pdf"
    )

    xlsx_name = (
        f"Laporan_Bulanan_"
        f"{gi_slug}_"
        f"{period_code}.xlsx"
    )

    pdf_bytes = build_monthly_pdf_bytes(
        bundle=bundle,
        report_year=report_year,
        report_month=report_month,
        gi_name=gi_name,
        gi_flc=gi_flc,
    )

    excel_bytes = build_monthly_excel_bytes(
        bundle=bundle,
        report_year=report_year,
        report_month=report_month,
        gi_name=gi_name,
    )

    pdf_result = (
        upload_monthly_report_bytes(
            monthly_report_id=(
                monthly_report_id
            ),
            report_year=(
                report_year
            ),
            report_month=(
                report_month
            ),
            gi_flc=(
                gi_flc
            ),
            gi_name=(
                gi_name
            ),
            file_name=(
                pdf_name
            ),
            mime_type=(
                "application/pdf"
            ),
            file_bytes=(
                pdf_bytes
            ),
            file_format=(
                "PDF"
            ),
        )
    )

    excel_result = (
        upload_monthly_report_bytes(
            monthly_report_id=(
                monthly_report_id
            ),
            report_year=(
                report_year
            ),
            report_month=(
                report_month
            ),
            gi_flc=(
                gi_flc
            ),
            gi_name=(
                gi_name
            ),
            file_name=(
                xlsx_name
            ),
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            file_bytes=(
                excel_bytes
            ),
            file_format=(
                "XLSX"
            ),
        )
    )

    return {
        "pdf":
            pdf_result,

        "xlsx":
            excel_result,
    }